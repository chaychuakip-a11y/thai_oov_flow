import pandas as pd
import sys
import os
import re
import pickle
from openpyxl import load_workbook

ILLEGEAL_CHAR = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]|\'')

def update_slots(file_path, slots):
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.strip(): continue
            if slots.get('new_oov_slots') is None:
                slots['new_oov_slots'] = {}
            slots['new_oov_slots'][line.strip().replace(' ', '')] = line.strip()
    return slots

def update_jushis(file_path, jushis):
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.strip(): continue
            if jushis.get(line.strip().replace(' ', '')) is None:
                jushis[line.strip().replace(' ', '')] = line.strip()
    return jushis

def remake_jushi_sheet(work_sheet, jushis_dict):
    data = []
    for row in work_sheet.iter_rows(values_only=True):
        new_row = []
        for cell in row:
            if cell is None or str(cell).strip() == '':
                new_row.append(cell)
                continue
            
            val_str = str(cell)
            cleaned_val = ILLEGEAL_CHAR.sub(r'', val_str)
            key = cleaned_val.strip().replace(' ', '')

            if jushis_dict.get(key) is not None:
                new_row.append(jushis_dict[key])
            else:
                new_row.append(cell)
        data.append(new_row)
    return data

def remake_slot_sheet(work_sheet, slots_dict):
    headers = [work_sheet.cell(row=1, column=col_idx).value for col_idx in range(1, work_sheet.max_column + 1)]
    data = []
    for row_idx, row in enumerate(work_sheet.iter_rows(values_only=True)):
        new_row = []
        for col_idx, cell in enumerate(row):
            if row_idx == 0:
                new_row.append(cell)
                continue
            if cell is None or str(cell).strip() == '':
                new_row.append(cell)
                continue

            slot_name = headers[col_idx]
            val_str = str(cell)
            cleaned_val = ILLEGEAL_CHAR.sub(r'', val_str)
            key = cleaned_val.strip().replace(' ', '')

            mapped_val = None
            if slot_name and slots_dict.get(slot_name) is not None and slots_dict[slot_name].get(key) is not None:
                mapped_val = slots_dict[slot_name][key]
            elif slots_dict.get('new_oov_slots') is not None and slots_dict['new_oov_slots'].get(key) is not None:
                mapped_val = slots_dict['new_oov_slots'][key]

            if mapped_val is not None:
                new_row.append(mapped_val)
            else:
                new_row.append(cell)
        data.append(new_row)
    return data

if __name__ == "__main__":
    temp_pkl_dir = sys.argv[1]
    res_dir = sys.argv[2]
    new_resources_dir = sys.argv[3]
    out_dir = sys.argv[4]
    memory_path = sys.argv[5]

    slots_dict = {'new_oov_slots': {}}
    jushis_dict = {}
    if os.path.exists(memory_path):
        with open(memory_path, 'rb') as f:
            memory = pickle.load(f)
            slots_dict = memory.get('slots', {'new_oov_slots': {}})
            jushis_dict = memory.get('jushis', {})
        
    for file in os.listdir(res_dir):
        if not file.endswith('.tts_out.split.checked'): continue
        file_path = os.path.join(res_dir, file)
        if 'jushi' in file:
            jushis_dict = update_jushis(file_path, jushis_dict)
        else:
            slots_dict = update_slots(file_path, slots_dict)

    os.makedirs(out_dir, exist_ok=True)
    for file in os.listdir(new_resources_dir):
        if not file.endswith('.xlsx') or file.startswith('~'): continue
        file_path = os.path.join(new_resources_dir, file)
        out_path = os.path.join(out_dir, file)
        
        workbook = load_workbook(filename=file_path)
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            for s_name in workbook.sheetnames:
                sheet_now = workbook[s_name]
                if s_name != '<>':
                    data_now = remake_jushi_sheet(sheet_now, jushis_dict)
                else:
                    data_now = remake_slot_sheet(sheet_now, slots_dict)
                pd.DataFrame(data_now).to_excel(writer, sheet_name=s_name, index=False, header=False)

    updated_memory = {'slots': slots_dict, 'jushis': jushis_dict}
    os.makedirs(os.path.dirname(memory_path), exist_ok=True)
    with open(memory_path, 'wb') as f:
        pickle.dump(updated_memory, f)