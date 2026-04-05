import os
import sys
import pickle
import re
from openpyxl import load_workbook

# 过滤 Excel 中的非法字符
ILLEGEAL_CHAR = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]|\'')

def is_En(char):
    """判断是否为英文字符或特定符号"""
    return ('a' <= char <= 'z') or ('A' <= char <= 'Z') or char in "<>'"

def remake_line(line):
    """Slot 文本规整：在英中交界处补空格"""
    line = line.strip()
    new_line = ''
    for p in line:
        if p == ' ':
            if new_line and is_En(new_line[-1]):
                new_line += p
            continue
        if new_line and is_En(p) != is_En(new_line[-1]) and new_line[-1] != ' ':
            new_line += ' '
        new_line += p
    return new_line

def remake_jushi_line(line):
    """Jushi 文本规整：处理占位符并在交界处补空格"""
    new_out = []
    for char in line:
        if char == '<':
            new_out.append(char)
        elif char == '>':
            if new_out: new_out[-1] += char
            new_out.append('')
        else:
            if not new_out: new_out.append('')
            new_out[-1] += char
            
    line_out = []
    for part in new_out:
        if not part: continue
        if '<' in part:
            part = part.replace('<', 'cdxa ').replace('>', ' cdxb')
        else:
            part = remake_line(part)
        line_out.append(part)
    return ' '.join(line_out).strip() + ' yjchen'

def load_base_memory(memory_path):
    """加载已有的 Master Memory"""
    if os.path.exists(memory_path):
        with open(memory_path, 'rb') as f:
            return pickle.load(f)
    return {'slots': {'new_oov_slots': {}}, 'jushis': {}}

def process_old_resources(old_dir):
    """
    [初始化工具] 从原始旧语料库扫描并建立 Memory
    (修复了 ReadOnlyWorksheet 的 iter_rows 兼容性)
    """
    slots = {'new_oov_slots': {}}
    jushis = {}
    if not os.path.exists(old_dir):
        return slots, jushis

    for file in os.listdir(old_dir):
        if file.startswith('~') or not file.endswith('.xlsx'):
            continue
        file_path = os.path.join(old_dir, file)
        # 开启 read_only 以防大文件内存溢出
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        for s_name in workbook.sheetnames:
            sheet = workbook[s_name]
            if s_name == '<>':
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    headers = next(rows_iter) # 第一行作为列名 (Slot Name)
                except StopIteration:
                    continue
                
                for row in rows_iter:
                    for col_idx, slot in enumerate(row):
                        if slot and str(slot).strip():
                            if col_idx >= len(headers): continue
                            slot_name = headers[col_idx]
                            if not slot_name: continue
                            if slot_name not in slots: slots[slot_name] = {}
                            cleaned = ILLEGEAL_CHAR.sub('', str(slot))
                            slots[slot_name][cleaned.strip().replace(' ', '')] = cleaned
            else:
                for row in sheet.iter_rows(values_only=True):
                    for jushi in row:
                        if jushi and str(jushi).strip():
                            cleaned = ILLEGEAL_CHAR.sub('', str(jushi))
                            jushis[cleaned.strip().replace(' ', '')] = cleaned
    return slots, jushis

def extract_oov(new_dir, memory):
    """
    [核心逻辑] 对比新语料与 Memory，提取增量 OOV
    (修复了 ReadOnlyWorksheet 的 iter_rows 兼容性)
    """
    slots = memory.get('slots', {})
    jushis = memory.get('jushis', {})
    oov_slot, oov_jushi = set(), set()
    
    if not os.path.exists(new_dir):
        print(f"Error: New resources dir {new_dir} not found.")
        return list(oov_slot), list(oov_jushi)

    for file in os.listdir(new_dir):
        if file.startswith('~') or not file.endswith('.xlsx'):
            continue
        file_path = os.path.join(new_dir, file)
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        
        for s_name in workbook.sheetnames:
            sheet = workbook[s_name]
            if s_name == '<>':
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    headers = next(rows_iter)
                except StopIteration:
                    continue

                for row in rows_iter:
                    for col_idx, slot in enumerate(row):
                        if slot and str(slot).strip():
                            if col_idx >= len(headers): continue
                            slot_name = headers[col_idx]
                            if not slot_name: continue
                            cleaned = ILLEGEAL_CHAR.sub('', str(slot)).strip()
                            key = cleaned.replace(' ', '')
                            # 如果该 Slot 分类下的词条在 Memory 中不存在，则记为 OOV
                            if slot_name not in slots or key not in slots[slot_name]:
                                oov_slot.add(cleaned)
            else:
                for row in sheet.iter_rows(values_only=True):
                    for jushi in row:
                        if jushi and str(jushi).strip():
                            cleaned = ILLEGEAL_CHAR.sub('', str(jushi)).strip()
                            key = cleaned.replace(' ', '')
                            # 如果该 Jushi 在 Memory 中不存在，则记为 OOV
                            if key not in jushis:
                                oov_jushi.add(cleaned)
    return list(oov_slot), list(oov_jushi)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python scripts/0_extract_and_format_crops_v2.py <new_dir> <out_crops_dir> <memory_path>")
        sys.exit(1)

    new_dir = sys.argv[1]
    out_crops_dir = sys.argv[2]
    memory_path = sys.argv[3]
    
    os.makedirs(out_crops_dir, exist_ok=True)
    
    # 1. 加载 Memory
    memory = load_base_memory(memory_path)
    
    # 2. 提取 OOV
    oov_slots_list, oov_jushis_list = extract_oov(new_dir, memory)
    
    # 3. 输出 desplit 文件供 TTS 预测
    if oov_slots_list:
        with open(os.path.join(out_crops_dir, 'oov_slot.txt.desplit'), 'w', encoding='utf-8') as f:
            f.writelines(remake_line(slot) + '\n' for slot in oov_slots_list)
            
    if oov_jushis_list:
        with open(os.path.join(out_crops_dir, 'oov_jushi.txt.desplit'), 'w', encoding='utf-8') as f:
            f.writelines(remake_jushi_line(jushi) + '\n' for jushi in oov_jushis_list)
            
    print(f"Extraction complete. OOV Slots: {len(oov_slots_list)}, OOV Jushis: {len(oov_jushis_list)}")